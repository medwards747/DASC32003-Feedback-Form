"""
generate_report.py
------------------
Standalone CLI script — no Streamlit required.

Pulls all submissions for a student's team from Supabase, then writes a
three-sheet Excel workbook:

  Sheet 1 "Score Summary"      — team averages per question (AVERAGE formulas)
  Sheet 2 "Raw Scores"         — one row per submission (source data for Sheet 1)
  Sheet 3 "Individual Feedback"— only the critiques written about this student

Usage
-----
    # Generate report for one student
    python generate_report.py "Alice Johnson"

    # Save to a specific folder
    python generate_report.py "Alice Johnson" --out reports/

    # List every registered student (to check exact name spelling)
    python generate_report.py --list-students

Secrets  (PROFESSOR ONLY)
------------------------
This script uses the SERVICE_ROLE key, which bypasses Supabase Row Level
Security and has full read access to all submissions.  Do NOT share this
key with students or deploy it inside the Streamlit app.

Credentials are read from the first source found:
  1. Environment variables  SUPABASE_URL  and  SUPABASE_SERVICE_KEY
  2. .streamlit/secrets.toml  key  SUPABASE_SERVICE_KEY
"""

from __future__ import annotations

import argparse
import os
import sys
import tomllib
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from supabase_manager import Student, Submission, SupabaseManager, Team

# ---------------------------------------------------------------------------
# Palette  (all RGB hex strings for openpyxl)
# ---------------------------------------------------------------------------
CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"   # white
CLR_SUBHDR_BG   = "D6E4F0"   # light blue
CLR_SUBHDR_FG   = "1F3864"
CLR_ROW_ALT     = "EBF5FB"   # alternating row tint
CLR_ACCENT      = "2E86C1"   # section-title blue
CLR_BORDER      = "AEB6BF"   # medium-grey border

FONT_MAIN  = "Arial"
FONT_TITLE = "Arial"

Q_LABELS: dict[str, str] = {
    "q1": "Q1 — Communication",
    "q2": "Q2 — Contribution",
    "q3": "Q3 — Subject Knowledge",
    "q4": "Q4 — Deliverable Quality",
    "q5": "Q5 — Would Work Again",
}
ALL_QKEYS = list(Q_LABELS.keys())  # ["q1", …, "q5"]


# ---------------------------------------------------------------------------
# Secrets loader
# ---------------------------------------------------------------------------

def _load_secrets() -> tuple[str, str]:
    """
    Return (SUPABASE_URL, SUPABASE_SERVICE_KEY) for the professor's report tool.

    The service_role key bypasses Row Level Security — it gives full read
    access to submissions.  Never expose it in the student-facing Streamlit app.

    Lookup order:
      1. Environment variables SUPABASE_URL + SUPABASE_SERVICE_KEY
      2. .streamlit/secrets.toml  →  SUPABASE_SERVICE_KEY
    """
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if url and key:
        return url, key

    toml_path = Path(".streamlit/secrets.toml")
    if toml_path.exists():
        with toml_path.open("rb") as fh:
            data = tomllib.load(fh)
        url = url or data.get("SUPABASE_URL")
        key = data.get("SUPABASE_SERVICE_KEY")
        if url and key:
            return url, key

    sys.exit(
        "❌  Could not find the Supabase service-role key.\n"
        "    Set environment variables SUPABASE_URL and SUPABASE_SERVICE_KEY,\n"
        "    or add SUPABASE_SERVICE_KEY to .streamlit/secrets.toml.\n"
        "    NOTE: Do NOT use the anon key here — it cannot read submissions."
    )


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    s = Side(style="thin", color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row: int, col: int, value: str, width: int = 0) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name=FONT_MAIN, bold=True, color=CLR_HEADER_FG, size=11)
    cell.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _subheader_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name=FONT_MAIN, bold=True, color=CLR_SUBHDR_FG, size=10)
    cell.fill      = PatternFill("solid", start_color=CLR_SUBHDR_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _thin_border()


def _data_cell(ws, row: int, col: int, value, alt_row: bool = False, centre: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name=FONT_MAIN, size=10)
    cell.fill      = PatternFill("solid", start_color=(CLR_ROW_ALT if alt_row else "FFFFFF"))
    cell.alignment = Alignment(
        horizontal="center" if centre else "left",
        vertical="top",
        wrap_text=True,
    )
    cell.border = _thin_border()


def _formula_cell(ws, row: int, col: int, formula: str, alt_row: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font      = Font(name=FONT_MAIN, size=10, color="000000")
    cell.fill      = PatternFill("solid", start_color=(CLR_ROW_ALT if alt_row else "FFFFFF"))
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.00"
    cell.border    = _thin_border()


def _title_block(ws, student_name: str, team: Team, n_reviews: int) -> None:
    """Write the student / team header at the top of any sheet."""
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value     = "DASC32003 — Peer Review Feedback Report"
    title.font      = Font(name=FONT_TITLE, bold=True, size=14, color=CLR_HEADER_FG)
    title.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("Student",        student_name),
        ("Team",           f"Team {team.team_number} — {team.team_name}"),
        ("Reviews Received", str(n_reviews)),
        ("Generated",      datetime.now().strftime("%d %b %Y %H:%M")),
    ]
    for i, (label, val) in enumerate(meta, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        lbl = ws[f"A{i}"]
        lbl.value     = label
        lbl.font      = Font(name=FONT_MAIN, bold=True, size=10, color=CLR_SUBHDR_FG)
        lbl.fill      = PatternFill("solid", start_color=CLR_SUBHDR_BG)
        lbl.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(f"C{i}:G{i}")
        v = ws[f"C{i}"]
        v.value     = val
        v.font      = Font(name=FONT_MAIN, size=10)
        v.alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_raw_scores(ws, submissions: list[Submission]) -> int:
    """
    Sheet 2 – Raw Scores.
    Returns the Excel row number of the last data row (for AVERAGE references).
    """
    ws.title = "Raw Scores"

    # ── column headers ──────────────────────────────────────────────────
    headers = ["Reviewer", "Submitted At"] + [Q_LABELS[k] for k in ALL_QKEYS]
    widths  = [22, 18] + [22] * 5
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 1, col, h, width=w)
    ws.row_dimensions[1].height = 36

    # ── data rows ────────────────────────────────────────────────────────
    for r_idx, sub in enumerate(submissions, start=2):
        alt = (r_idx % 2 == 0)
        dt_str = sub.created_at[:16].replace("T", " ") if sub.created_at else ""
        _data_cell(ws, r_idx, 1, sub.reviewer_name, alt)
        _data_cell(ws, r_idx, 2, dt_str, alt, centre=True)
        for c_idx, qkey in enumerate(ALL_QKEYS, start=3):
            score = sub.scores.get(qkey)
            _data_cell(ws, r_idx, c_idx, score, alt, centre=True)

    ws.freeze_panes = "A2"
    return max(2, 1 + len(submissions))   # last data row


def _build_individual_feedback(ws, student_name: str, submissions: list[Submission]) -> None:
    """Sheet 3 – Individual Feedback (only comments about this student)."""
    ws.title = "Individual Feedback"

    headers = ["Reviewer", "Submitted At", "Feedback"]
    widths  = [22, 18, 70]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_cell(ws, 1, col, h, width=w)
    ws.row_dimensions[1].height = 36

    row = 2
    for sub in submissions:
        feedback_text = sub.individual_feedback.get(student_name, "").strip()
        if not feedback_text:
            continue
        alt = (row % 2 == 0)
        dt_str = sub.created_at[:16].replace("T", " ") if sub.created_at else ""
        _data_cell(ws, row, 1, sub.reviewer_name, alt)
        _data_cell(ws, row, 2, dt_str, alt, centre=True)
        _data_cell(ws, row, 3, feedback_text, alt)
        ws.row_dimensions[row].height = max(
            30, min(15 * (1 + len(feedback_text) // 80), 120)
        )
        row += 1

    if row == 2:
        ws.merge_cells("A2:C2")
        cell = ws["A2"]
        cell.value     = "No individual feedback was recorded for this student."
        cell.font      = Font(name=FONT_MAIN, italic=True, color="808080", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"


def _build_summary(
    ws,
    student_name: str,
    team: Team,
    submissions: list[Submission],
    last_data_row: int,
) -> None:
    """Sheet 1 – Score Summary with AVERAGE() formulas referencing 'Raw Scores'."""
    ws.title = "Score Summary"

    _title_block(ws, student_name, team, len(submissions))

    # ── section heading ───────────────────────────────────────────────────
    START_ROW = 7
    ws.merge_cells(f"A{START_ROW}:E{START_ROW}")
    sec = ws[f"A{START_ROW}"]
    sec.value     = "Team Average Scores (all reviewers)"
    sec.font      = Font(name=FONT_MAIN, bold=True, size=11, color=CLR_ACCENT)
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[START_ROW].height = 22

    # ── table headers ─────────────────────────────────────────────────────
    TBL_HDR = START_ROW + 1
    col_headers = ["Question", "Label", "Average (1–5)", "Likert Descriptor"]
    col_widths  = [14, 38, 16, 26]
    for col, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        _header_cell(ws, TBL_HDR, col, h, width=w)
    ws.row_dimensions[TBL_HDR].height = 30

    # ── one row per question, AVERAGE formula referencing Raw Scores ──────
    DESCRIPTORS = {
        "q1": "Q1 — Communication",
        "q2": "Q2 — Contribution",
        "q3": "Q3 — Subject Knowledge",
        "q4": "Q4 — Deliverable Quality",
        "q5": "Q5 — Would Work Again",
    }
    # Raw Scores columns: Reviewer=A, Date=B, Q1=C, Q2=D, Q3=E, Q4=F, Q5=G
    RAW_SCORE_COLS = {"q1": "C", "q2": "D", "q3": "E", "q4": "F", "q5": "G"}

    avg_formula_rows: list[str] = []  # track cell addresses for overall avg

    for i, qkey in enumerate(ALL_QKEYS):
        r = TBL_HDR + 1 + i
        alt = (i % 2 == 0)
        rs_col = RAW_SCORE_COLS[qkey]
        # IFERROR guards against #DIV/0! when there are no submissions
        avg_formula = (
            f"=IFERROR(AVERAGE('Raw Scores'!{rs_col}2:{rs_col}{last_data_row}),\"—\")"
        )
        _subheader_cell(ws, r, 1, qkey.upper())
        _data_cell(ws, r, 2, DESCRIPTORS[qkey], alt)
        _formula_cell(ws, r, 3, avg_formula, alt)
        likert_desc = (
            f'=IFERROR(CHOOSE(ROUND(C{r},0),'
            '"Strongly Disagree","Disagree","Neutral","Agree","Strongly Agree"),"—")'
        )
        cell = ws.cell(row=r, column=4, value=likert_desc)
        cell.font      = Font(name=FONT_MAIN, size=10, color="000000")
        cell.fill      = PatternFill("solid", start_color=(CLR_ROW_ALT if alt else "FFFFFF"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        avg_formula_rows.append(f"C{r}")

    # ── overall average row ───────────────────────────────────────────────
    OVRL_ROW = TBL_HDR + 1 + len(ALL_QKEYS)
    overall_range = ",".join(avg_formula_rows)
    ws.merge_cells(f"A{OVRL_ROW}:B{OVRL_ROW}")
    ovrl_lbl = ws[f"A{OVRL_ROW}"]
    ovrl_lbl.value     = "Overall Average"
    ovrl_lbl.font      = Font(name=FONT_MAIN, bold=True, size=10, color=CLR_HEADER_FG)
    ovrl_lbl.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    ovrl_lbl.alignment = Alignment(horizontal="right", vertical="center")
    ovrl_lbl.border    = _thin_border()

    _formula_cell(ws, OVRL_ROW, 3, f"=IFERROR(AVERAGE({overall_range}),\"—\")")
    ovrl_c4 = ws.cell(row=OVRL_ROW, column=4)
    ovrl_c4.font   = Font(name=FONT_MAIN, bold=True, size=10, color=CLR_HEADER_FG)
    ovrl_c4.fill   = PatternFill("solid", start_color=CLR_HEADER_BG)
    ovrl_c4.border = _thin_border()

    ws.row_dimensions[OVRL_ROW].height = 22
    ws.freeze_panes = "A7"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_report(student_name: str, out_dir: str = ".") -> Path:
    url, key = _load_secrets()
    db = SupabaseManager(url=url, key=key)

    # Resolve student → team
    student: Student | None = db.get_student_by_name(student_name)
    if student is None:
        sys.exit(f"❌  No student named '{student_name}' found in the database.")

    teams = db.get_teams()
    team_map: dict[int, Team] = {t.id: t for t in teams}
    team = team_map.get(student.team_id)
    if team is None:
        sys.exit(f"❌  Team ID {student.team_id} not found.")

    # Fetch all submissions for that team
    submissions: list[Submission] = db.get_submissions_by_team(team.id)

    # Build workbook
    wb = Workbook()
    ws_summary  = wb.active
    ws_raw      = wb.create_sheet()
    ws_feedback = wb.create_sheet()

    last_data_row = _build_raw_scores(ws_raw, submissions)
    _build_individual_feedback(ws_feedback, student_name, submissions)
    _build_summary(ws_summary, student_name, team, submissions, last_data_row)

    # Reorder sheets: Summary first
    wb.move_sheet("Score Summary", offset=-wb.sheetnames.index("Score Summary"))

    # Save
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    safe_name = student_name.replace(" ", "_").replace("/", "-")
    filename  = out_path / f"feedback_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename


def _list_students() -> None:
    url, key = _load_secrets()
    db = SupabaseManager(url=url, key=key)
    students = db.get_all_students()
    teams    = {t.id: t for t in db.get_teams()}
    if not students:
        print("No students found.")
        return
    print(f"{'Name':<30}  {'Team'}")
    print("-" * 55)
    for s in students:
        t = teams.get(s.team_id)
        team_label = f"Team {t.team_number} — {t.team_name}" if t else f"team_id={s.team_id}"
        print(f"{s.name:<30}  {team_label}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate a per-student Excel feedback report from Supabase."
    )
    parser.add_argument(
        "student_name",
        nargs="?",
        help='Full name of the student (e.g. "Alice Johnson").',
    )
    parser.add_argument(
        "--out",
        default=".",
        metavar="DIR",
        help="Directory in which to save the .xlsx file (default: current directory).",
    )
    parser.add_argument(
        "--list-students",
        action="store_true",
        help="Print all registered students and exit.",
    )
    args = parser.parse_args()

    if args.list_students:
        _list_students()
        sys.exit(0)

    if not args.student_name:
        parser.error('Provide a student name or use --list-students.')

    output_file = generate_report(args.student_name, args.out)
    print(f"✅  Report saved → {output_file}")
