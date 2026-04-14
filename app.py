"""
app.py  —  DASC32003 Feedback System
Three tabs:
  1. Presentation Feedback  (students review other teams)
  2. Peer Reviews           (students review own teammates)
  3. Download Results       (professor downloads per-student Excel reports)
"""

from __future__ import annotations

import io
from datetime import datetime

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from supabase_manager import (
    PeerReviewPayload,
    Student,
    Submission,
    PeerReview,
    SubmissionPayload,
    SupabaseManager,
    Team,
)

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DASC32003 Feedback",
    page_icon="📋",
    layout="centered",
)

# ── Survey constants ───────────────────────────────────────────────────────
PRES_QUESTIONS: list[str] = [
    "The team provided the motivating information required for me to appreciate the problem's importance.",
    "The team introduced the problem in a way that I could understand.",
    "I understand the overall idea of the optimization model.",
    "I understand the mathematical formulation of the optimization model.",
    "I understand how the optimization model's parameters were populated using data.",
    "The analysis of optimization results was compelling.",
    "The slides are high quality and helpful for understanding the project.",
    "The presentation was well-organized with effective transitions.",
    "The presentation was well-rehearsed but extemporaneous.",
    "The team came across as confident, enthusiastic and knowledgeable about the project.",
]

PEER_QUESTION_TEMPLATES: list[str] = [
    "Please rate {name}'s contributions to generating ideas and planning the project.",
    "Please rate {name}'s contributions to identifying/gathering/analyzing data for the project.",
    "Please rate {name}'s contributions to formulating the optimization model.",
    "Please rate {name}'s contributions to implementing and solving the optimization model.",
    "Please rate {name}'s contributions to analyzing model output.",
    "Please rate {name}'s contributions to preparing the presentation.",
]

PEER_SCALE_LABELS: dict[int, str] = {
    1: "1 — Detrimental to the group",
    2: "2 — Adequate",
    3: "3 — Critical to the group's success",
}

# ── DB initialisation (once per session) ───────────────────────────────────

def _make_db(key_secret: str) -> SupabaseManager | None:
    try:
        return SupabaseManager(
            url=st.secrets["SUPABASE_URL"],
            key=st.secrets[key_secret],
        )
    except KeyError as exc:
        st.error(f"🔑 Missing secret **{exc}** in secrets.toml.")
        return None
    except Exception as exc:
        st.error(f"🔌 Database connection failed: `{exc}`")
        return None


if "db_anon" not in st.session_state:
    st.session_state["db_anon"] = _make_db("SUPABASE_ANON_KEY")

db: SupabaseManager | None = st.session_state["db_anon"]

if db is None:
    st.stop()

# ── Cached/session-cached data fetchers ───────────────────────────────────

def fetch_teams() -> list[Team]:
    if "_all_teams" not in st.session_state:
        try:
            st.session_state["_all_teams"] = db.get_teams()  # type: ignore[union-attr]
        except RuntimeError as exc:
            st.error(str(exc))
            st.session_state["_all_teams"] = []
    return st.session_state["_all_teams"]


def fetch_team_students(team_id: int) -> list[Student]:
    key = f"_students_{team_id}"
    if key not in st.session_state:
        try:
            st.session_state[key] = db.get_students_by_team(team_id)  # type: ignore[union-attr]
        except RuntimeError as exc:
            st.error(str(exc))
            st.session_state[key] = []
    return st.session_state[key]


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — PRESENTATION FEEDBACK
# ══════════════════════════════════════════════════════════════════════════════

def _render_presentation_tab() -> None:
    st.header("📊 Presentation Feedback")
    st.caption("Rate each team you observe. You may not review your own team.")

    # ── Step 1: Student ID verification ─────────────────────────────────
    col_inp, col_btn = st.columns([3, 1])
    sid_raw = col_inp.text_input("Enter your Student ID", key="t1_sid_input",
                                  placeholder="e.g. 483920571")
    verify = col_btn.button("Verify", key="t1_verify", use_container_width=True)

    if verify:
        st.session_state["t1_student"] = None
        if sid_raw.strip():
            try:
                sid = int(sid_raw.strip())
                student = db.get_student_by_random_id(sid)  # type: ignore[union-attr]
                if student:
                    st.session_state["t1_student"] = student
                else:
                    st.error("❌ Student ID not found. Check the number and try again.")
            except ValueError:
                st.error("❌ Please enter a numeric Student ID.")
            except RuntimeError as exc:
                st.error(str(exc))

    student: Student | None = st.session_state.get("t1_student")

    if not student:
        return

    st.success(f"✅ Verified: **{student.name}**")

    # ── Step 2: Team selection (exclude own team) ────────────────────────
    teams = fetch_teams()
    other_teams = [t for t in teams if t.id != student.team_id]
    if not other_teams:
        st.warning("No other teams are registered yet.")
        return

    team_map: dict[str, Team] = {f"Team {t.team_number}": t for t in other_teams}
    choice = st.selectbox("Select the team you are reviewing",
                          ["— please select —"] + list(team_map.keys()),
                          key="t1_team_choice")

    if choice == "— please select —":
        return

    selected_team = team_map[choice]
    team_students = fetch_team_students(selected_team.id)

    # ── Step 3: Feedback form ────────────────────────────────────────────
    if st.session_state.get("t1_success"):
        st.success("🎉 Your feedback was submitted successfully!")
        if st.button("Submit feedback for another team", key="t1_reset"):
            st.session_state["t1_success"] = False
            st.session_state["t1_student"] = None
            st.rerun()
        return

    with st.form("presentation_form"):
        st.subheader("Team Assessment")
        st.caption("1 = Strongly Disagree → 5 = Strongly Agree")

        scores: dict[str, int | None] = {}
        for idx, question in enumerate(PRES_QUESTIONS, start=1):
            scores[f"q{idx}"] = st.radio(
                f"**{idx}.** {question}",
                options=[1, 2, 3, 4, 5],
                format_func=lambda v: str(v),
                index=None,
                horizontal=True,
                key=f"t1_q{idx}",
            )

        st.subheader("Individual Feedback")
        st.caption("Leave specific comments for each team member.")
        ind_feedback: dict[str, str] = {}
        for s in team_students:
            ind_feedback[str(s.student_id)] = st.text_area(
                f"Feedback for **{s.name}**",
                placeholder=f"Write your feedback for {s.name} here…",
                height=90,
                key=f"t1_fb_{s.id}",
            )

        submitted = st.form_submit_button("Submit Feedback ✅",
                                          use_container_width=True, type="primary")

    if submitted:
        unanswered = [f"Q{i}" for i, k in enumerate(scores, start=1)
                      if scores[f"q{i}"] is None]
        if unanswered:
            st.error(f"Please answer all questions. Unanswered: {', '.join(unanswered)}")
            return

        payload = SubmissionPayload(
            submitter_student_id=student.student_id,
            reviewed_team_id=selected_team.id,
            scores={k: v for k, v in scores.items()},  # type: ignore[misc]
            individual_feedback=ind_feedback,
        )
        try:
            db.insert_submission(payload)  # type: ignore[union-attr]
            st.session_state["t1_success"] = True
            st.rerun()
        except RuntimeError as exc:
            msg = str(exc)
            if "uq_submission_student_team" in msg:
                st.error("⚠️ You have already submitted feedback for this team.")
            else:
                st.error(f"Submission failed: {msg}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — PEER REVIEWS
# ══════════════════════════════════════════════════════════════════════════════

def _render_peer_review_tab() -> None:
    st.header("👥 Peer Reviews")
    st.caption("Rate each of your teammates on their contributions to the project.")

    # ── Step 1: Student ID verification ─────────────────────────────────
    col_inp, col_btn = st.columns([3, 1])
    sid_raw = col_inp.text_input("Enter your Student ID", key="t2_sid_input",
                                  placeholder="e.g. 483920571")
    verify = col_btn.button("Verify", key="t2_verify", use_container_width=True)

    if verify:
        st.session_state["t2_student"] = None
        if sid_raw.strip():
            try:
                sid = int(sid_raw.strip())
                student = db.get_student_by_random_id(sid)  # type: ignore[union-attr]
                if student:
                    st.session_state["t2_student"] = student
                else:
                    st.error("❌ Student ID not found.")
            except ValueError:
                st.error("❌ Please enter a numeric Student ID.")
            except RuntimeError as exc:
                st.error(str(exc))

    student: Student | None = st.session_state.get("t2_student")
    if not student:
        return

    st.success(f"✅ Verified: **{student.name}**")

    teammates = [s for s in fetch_team_students(student.team_id) if s.id != student.id]
    if not teammates:
        st.warning("No teammates found for your team.")
        return

    # ── Step 2: Peer review form ─────────────────────────────────────────
    if st.session_state.get("t2_success"):
        st.success("🎉 Your peer reviews were submitted successfully!")
        if st.button("Done", key="t2_reset"):
            st.session_state["t2_success"] = False
            st.session_state["t2_student"] = None
            st.rerun()
        return

    with st.form("peer_review_form"):
        all_scores: dict[int, dict[str, int | None]] = {}
        all_texts:  dict[int, str] = {}

        for teammate in teammates:
            st.subheader(f"📌 {teammate.name}")
            tm_scores: dict[str, int | None] = {}
            for idx, template in enumerate(PEER_QUESTION_TEMPLATES, start=1):
                question = template.format(name=teammate.name.split()[0])
                tm_scores[f"q{idx}"] = st.radio(
                    question,
                    options=[1, 2, 3],
                    format_func=lambda v: PEER_SCALE_LABELS[v],
                    index=None,
                    horizontal=True,
                    key=f"t2_{teammate.id}_q{idx}",
                )
            all_texts[teammate.id] = st.text_area(
                f"Explain your evaluation of {teammate.name.split()[0]}",
                placeholder="Provide specific comments supporting your ratings above…",
                height=90,
                key=f"t2_{teammate.id}_text",
            )
            all_scores[teammate.id] = tm_scores
            st.divider()

        submitted = st.form_submit_button("Submit Peer Reviews ✅",
                                          use_container_width=True, type="primary")

    if submitted:
        errors: list[str] = []
        for teammate in teammates:
            unanswered = [f"Q{i}" for i in range(1, 7)
                          if all_scores[teammate.id].get(f"q{i}") is None]
            if unanswered:
                errors.append(
                    f"{teammate.name}: unanswered {', '.join(unanswered)}"
                )
        if errors:
            for e in errors:
                st.error(f"❌ {e}")
            return

        failed: list[str] = []
        for teammate in teammates:
            payload = PeerReviewPayload(
                submitter_student_id=student.student_id,
                reviewee_student_id=teammate.student_id,
                scores={k: v for k, v in all_scores[teammate.id].items()},  # type: ignore[misc]
                feedback_text=all_texts[teammate.id],
            )
            try:
                db.insert_peer_review(payload)  # type: ignore[union-attr]
            except RuntimeError as exc:
                msg = str(exc)
                if "uq_peer_review_pair" in msg:
                    failed.append(f"{teammate.name}: already reviewed")
                else:
                    failed.append(f"{teammate.name}: {msg}")

        if failed:
            for f in failed:
                st.error(f"⚠️ {f}")
        else:
            st.session_state["t2_success"] = True
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — DOWNLOAD RESULTS  (professor only, requires service-role key)
# ══════════════════════════════════════════════════════════════════════════════

# ── Excel style helpers ────────────────────────────────────────────────────

_NAV = "1F3864"
_WHT = "FFFFFF"
_BLU = "D6E4F0"
_ALT = "EBF5FB"
_BDR = "AEB6BF"
_FNT = "Arial"


def _border() -> Border:
    s = Side(style="thin", color=_BDR)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_cell(ws, r: int, c: int, v: str, w: int = 0) -> None:
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name=_FNT, bold=True, color=_WHT, size=10)
    cell.fill      = PatternFill("solid", start_color=_NAV)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()
    if w:
        ws.column_dimensions[get_column_letter(c)].width = w


def _dat_cell(ws, r: int, c: int, v, alt: bool = False, ctr: bool = False) -> None:
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name=_FNT, size=10)
    cell.fill      = PatternFill("solid", start_color=(_ALT if alt else _WHT))
    cell.alignment = Alignment(
        horizontal="center" if ctr else "left",
        vertical="top",
        wrap_text=True,
    )
    cell.border = _border()


def _title_block(ws, student: Student, team: Team, n_pres: int, n_peer: int) -> None:
    for merged in ["A1:K1", "A2:K2", "A3:K3", "A4:K4", "A5:K5"]:
        ws.merge_cells(merged)

    def _meta(row: int, label: str, value: str) -> None:
        c = ws.cell(row=row, column=1, value=f"{label}:  {value}")
        c.font      = Font(name=_FNT, size=10,
                           bold=(row == 1),
                           color=(_WHT if row == 1 else _NAV))
        c.fill      = PatternFill("solid", start_color=(_NAV if row == 1 else _BLU))
        c.alignment = Alignment(horizontal="center" if row == 1 else "left",
                                vertical="center")

    _meta(1, "DASC32003 Peer-Review Report", student.name)
    _meta(2, "Team", f"Team {team.team_number}")
    _meta(3, "Presentation reviews received", str(n_pres))
    _meta(4, "Peer reviews received", str(n_peer))
    _meta(5, "Generated", datetime.now().strftime("%d %b %Y %H:%M"))
    ws.row_dimensions[1].height = 26


def _build_team_performance_sheet(
    ws,
    student: Student,
    team: Team,
    submissions: list[Submission],
    peer_reviews_for_student: list[PeerReview],
) -> None:
    ws.title = "Team Performance"

    _title_block(ws, student, team, len(submissions), len(peer_reviews_for_student))

    # ── Score-summary section ────────────────────────────────────────────
    SUMM_HDR = 7
    ws.merge_cells(f"A{SUMM_HDR}:K{SUMM_HDR}")
    sec = ws.cell(row=SUMM_HDR, column=1,
                  value="Presentation Feedback — Team Averages (all reviewers)")
    sec.font      = Font(name=_FNT, bold=True, size=11, color=_NAV)
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[SUMM_HDR].height = 20

    TBL_HDR = SUMM_HDR + 1
    for col, (label, w) in enumerate(
        [("Q#", 5), ("Question", 60), ("Team Average (1–5)", 20)], start=1
    ):
        _hdr_cell(ws, TBL_HDR, col, label, w)
    ws.row_dimensions[TBL_HDR].height = 30

    PRES_Q_LABELS = [
        "Motivating information",
        "Problem introduction",
        "Overall model idea",
        "Mathematical formulation",
        "Parameter population",
        "Results analysis",
        "Slide quality",
        "Organisation & transitions",
        "Rehearsal & delivery",
        "Confidence & enthusiasm",
    ]

    # Raw data will be written starting at RAW_DATA_ROW
    RAW_HDR_ROW  = TBL_HDR + len(PRES_QUESTIONS) + 3
    RAW_DATA_ROW = RAW_HDR_ROW + 1
    raw_last_row = RAW_DATA_ROW + len(submissions) - 1 if submissions else RAW_DATA_ROW

    for i, label in enumerate(PRES_Q_LABELS, start=1):
        r   = TBL_HDR + i
        alt = (i % 2 == 0)
        # Raw data: col A = row#, cols B–K = Q1–Q10
        raw_col = get_column_letter(i + 1)   # B=Q1, C=Q2 … K=Q10
        formula = (
            f"=IFERROR(AVERAGE({raw_col}{RAW_DATA_ROW}:{raw_col}{raw_last_row}),\"—\")"
        )
        _dat_cell(ws, r, 1, f"Q{i}", alt, ctr=True)
        _dat_cell(ws, r, 2, label, alt)
        avg_cell = ws.cell(row=r, column=3, value=formula)
        avg_cell.font         = Font(name=_FNT, size=10)
        avg_cell.fill         = PatternFill("solid", start_color=(_ALT if alt else _WHT))
        avg_cell.alignment    = Alignment(horizontal="center", vertical="center")
        avg_cell.number_format = "0.00"
        avg_cell.border       = _border()

    # Overall average row
    OVRL = TBL_HDR + len(PRES_QUESTIONS) + 1
    ws.merge_cells(f"A{OVRL}:B{OVRL}")
    lbl = ws.cell(row=OVRL, column=1, value="Overall Average")
    lbl.font      = Font(name=_FNT, bold=True, color=_WHT, size=10)
    lbl.fill      = PatternFill("solid", start_color=_NAV)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border    = _border()

    ovrl_range = ",".join(f"C{TBL_HDR + i}" for i in range(1, 11))
    oc = ws.cell(row=OVRL, column=3,
                 value=f"=IFERROR(AVERAGE({ovrl_range}),\"—\")")
    oc.font         = Font(name=_FNT, bold=True, color=_WHT, size=10)
    oc.fill         = PatternFill("solid", start_color=_NAV)
    oc.alignment    = Alignment(horizontal="center", vertical="center")
    oc.number_format = "0.00"
    oc.border       = _border()

    # ── Raw submissions data ─────────────────────────────────────────────
    ws.merge_cells(f"A{RAW_HDR_ROW - 1}:K{RAW_HDR_ROW - 1}")
    raw_sec = ws.cell(row=RAW_HDR_ROW - 1, column=1,
                      value="Raw Submission Data (source for averages above)")
    raw_sec.font      = Font(name=_FNT, bold=True, size=10, color=_NAV)
    raw_sec.alignment = Alignment(horizontal="left", vertical="center")

    raw_headers = ["#", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9", "Q10"]
    for col, h in enumerate(raw_headers, start=1):
        _hdr_cell(ws, RAW_HDR_ROW, col, h)
    ws.row_dimensions[RAW_HDR_ROW].height = 22

    for row_idx, sub in enumerate(submissions, start=1):
        r   = RAW_DATA_ROW + row_idx - 1
        alt = (row_idx % 2 == 0)
        _dat_cell(ws, r, 1, row_idx, alt, ctr=True)
        for qi in range(1, 11):
            _dat_cell(ws, r, qi + 1, sub.scores.get(f"q{qi}"), alt, ctr=True)

    ws.freeze_panes = "A8"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20


def _build_individual_feedback_sheet(
    ws,
    student: Student,
    submissions: list[Submission],
) -> None:
    ws.title = "Individual Feedback"

    ws.merge_cells("A1:C1")
    h = ws.cell(row=1, column=1,
                value=f"Individual Feedback for {student.name} (reviewer names redacted)")
    h.font      = Font(name=_FNT, bold=True, color=_WHT, size=11)
    h.fill      = PatternFill("solid", start_color=_NAV)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    _hdr_cell(ws, 2, 1, "#",            w=5)
    _hdr_cell(ws, 2, 2, "Submitted At", w=18)
    _hdr_cell(ws, 2, 3, "Feedback",     w=80)
    ws.row_dimensions[2].height = 26

    sid_key = str(student.student_id)
    row_num = 3
    for sub in submissions:
        text = sub.individual_feedback.get(sid_key, "").strip()
        if not text:
            continue
        alt = (row_num % 2 == 0)
        dt  = sub.created_at[:16].replace("T", " ") if sub.created_at else ""
        _dat_cell(ws, row_num, 1, row_num - 2, alt, ctr=True)
        _dat_cell(ws, row_num, 2, dt, alt, ctr=True)
        _dat_cell(ws, row_num, 3, text, alt)
        ws.row_dimensions[row_num].height = max(
            30, min(15 * (1 + len(text) // 80), 120)
        )
        row_num += 1

    if row_num == 3:
        ws.merge_cells("A3:C3")
        nc = ws.cell(row=3, column=1, value="No individual feedback recorded yet.")
        nc.font      = Font(name=_FNT, italic=True, color="808080", size=10)
        nc.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"


def _build_peer_reviews_sheet(
    ws,
    student: Student,
    peer_reviews: list[PeerReview],
) -> None:
    ws.title = "Peer Reviews"

    ws.merge_cells("A1:J1")
    h = ws.cell(row=1, column=1,
                value=f"Peer Reviews for {student.name} (reviewer names redacted)")
    h.font      = Font(name=_FNT, bold=True, color=_WHT, size=11)
    h.fill      = PatternFill("solid", start_color=_NAV)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["#", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Avg", "Comments"]
    widths  = [4,    8,    8,    8,    8,    8,    8,    8,    60]
    for col, (label, w) in enumerate(zip(headers, widths), start=1):
        _hdr_cell(ws, 2, col, label, w)
    ws.row_dimensions[2].height = 26

    for row_idx, pr in enumerate(peer_reviews, start=1):
        r   = row_idx + 2
        alt = (row_idx % 2 == 0)
        scores = [pr.scores.get(f"q{i}") for i in range(1, 7)]
        avg    = sum(s for s in scores if s) / max(len([s for s in scores if s]), 1)
        _dat_cell(ws, r, 1, row_idx, alt, ctr=True)
        for ci, sc in enumerate(scores, start=2):
            _dat_cell(ws, r, ci, sc, alt, ctr=True)
        avg_c = ws.cell(row=r, column=8,
                        value=f"=IFERROR(AVERAGE(B{r}:G{r}),\"—\")")
        avg_c.font         = Font(name=_FNT, size=10)
        avg_c.fill         = PatternFill("solid", start_color=(_ALT if alt else _WHT))
        avg_c.alignment    = Alignment(horizontal="center", vertical="center")
        avg_c.number_format = "0.00"
        avg_c.border       = _border()
        _dat_cell(ws, r, 9, pr.feedback_text, alt)
        ws.row_dimensions[r].height = max(
            25, min(15 * (1 + len(pr.feedback_text) // 80), 120)
        )

    if not peer_reviews:
        ws.merge_cells("A3:J3")
        nc = ws.cell(row=3, column=1, value="No peer reviews recorded yet.")
        nc.font      = Font(name=_FNT, italic=True, color="808080", size=10)
        nc.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"


def _generate_excel_bytes(
    student: Student,
    team: Team,
    all_submissions: list[Submission],
    all_peer_reviews: list[PeerReview],
) -> bytes:
    team_subs      = [s for s in all_submissions if s.reviewed_team_id == team.id]
    student_peers  = [pr for pr in all_peer_reviews
                      if pr.reviewee_student_id == student.student_id]

    wb = Workbook()
    ws1 = wb.active
    _build_team_performance_sheet(ws1, student, team, team_subs, student_peers)
    ws2 = wb.create_sheet()
    _build_individual_feedback_sheet(ws2, student, team_subs)
    ws3 = wb.create_sheet()
    _build_peer_reviews_sheet(ws3, student, student_peers)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_download_tab() -> None:
    st.header("📥 Download Results")
    st.caption("Professor access only. Generates one Excel workbook per student.")

    if not st.secrets.get("SUPABASE_SERVICE_KEY"):
        st.warning(
            "⚠️ `SUPABASE_SERVICE_KEY` is not configured. "
            "This tab only works when running locally with a full secrets.toml."
        )
        return

    col_inp, col_btn = st.columns([3, 1])
    prof_id_raw = col_inp.text_input("Professor ID", type="password",
                                     key="t3_prof_id")
    auth_click  = col_btn.button("Authenticate", key="t3_auth",
                                 use_container_width=True)

    if auth_click:
        expected = str(st.secrets.get("PROFESSOR_ID", ""))
        if prof_id_raw.strip() == expected and expected:
            st.session_state["t3_authed"] = True
            # Clear cached data so a fresh load happens
            for k in ["t3_students", "t3_teams", "t3_submissions", "t3_peer_reviews"]:
                st.session_state.pop(k, None)
        else:
            st.session_state["t3_authed"] = False
            st.error("❌ Invalid Professor ID.")

    if not st.session_state.get("t3_authed"):
        return

    # ── Load all data (service role) ─────────────────────────────────────
    if "db_service" not in st.session_state:
        st.session_state["db_service"] = _make_db("SUPABASE_SERVICE_KEY")

    db_svc: SupabaseManager | None = st.session_state["db_service"]
    if db_svc is None:
        return

    with st.spinner("Loading data…"):
        if "t3_students" not in st.session_state:
            st.session_state["t3_students"]    = db_svc.get_all_students()
            st.session_state["t3_teams"]       = {t.id: t for t in db_svc.get_teams()}
            st.session_state["t3_submissions"] = db_svc.get_all_submissions()
            st.session_state["t3_peer_reviews"]= db_svc.get_all_peer_reviews()

    students:     list[Student]         = st.session_state["t3_students"]
    teams:        dict[int, Team]       = st.session_state["t3_teams"]
    all_subs:     list[Submission]      = st.session_state["t3_submissions"]
    all_peers:    list[PeerReview]      = st.session_state["t3_peer_reviews"]

    if not students:
        st.info("No students in the database yet.")
        return

    st.success(f"✅ Authenticated. {len(students)} students · "
               f"{len(all_subs)} presentation submissions · "
               f"{len(all_peers)} peer reviews")
    st.divider()

    current_team_id = None
    for student in students:
        team = teams.get(student.team_id)
        if team is None:
            continue
        if team.id != current_team_id:
            st.subheader(f"Team {team.team_number}")
            current_team_id = team.id

        col_name, col_btn = st.columns([3, 1])
        col_name.write(student.name)
        excel_bytes = _generate_excel_bytes(student, team, all_subs, all_peers)
        col_btn.download_button(
            label="📥 Download",
            data=excel_bytes,
            file_name=f"feedback_{student.name.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"t3_dl_{student.id}",
            use_container_width=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# Main layout
# ══════════════════════════════════════════════════════════════════════════════

st.title("📋 DASC32003 Feedback System")

tab1, tab2, tab3 = st.tabs(
    ["📊 Presentation Feedback", "👥 Peer Reviews", "📥 Download Results"]
)

with tab1:
    _render_presentation_tab()

with tab2:
    _render_peer_review_tab()

with tab3:
    _render_download_tab()
